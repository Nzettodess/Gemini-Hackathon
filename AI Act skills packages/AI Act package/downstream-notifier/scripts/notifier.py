import os
import sys
import argparse
import subprocess
import json
import sqlite3
import hashlib
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl
import shutil
from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parent.parent.parent.parent.parent / ".env")

# Try to import Gemini client
try:
    from google import genai
except ImportError:
    genai = None

class SnapshotDiff:
    """Handles diffing local directories using SQLite hashing with concurrency protection."""
    def __init__(self, db_path="examples/notices.db"):
        self.db_path = db_path
        self._init_db()

    def _init_db(self):
        # Added timeout for basic concurrent access protection
        with sqlite3.connect(self.db_path, timeout=30) as conn:
            conn.execute("CREATE TABLE IF NOT EXISTS snapshots (file_path TEXT PRIMARY KEY, hash TEXT, last_seen TIMESTAMP)")

    def get_diff(self, directory):
        """Detect changes in a directory based on file hashes."""
        changes = []
        current_files = set()
        
        with sqlite3.connect(self.db_path, timeout=30) as conn:
            for root, _, files in os.walk(directory):
                if ".git" in root: continue
                for file in files:
                    full_path = Path(root) / file
                    current_files.add(str(full_path))
                    try:
                        content = full_path.read_bytes()
                        file_hash = hashlib.sha256(content).hexdigest()
                        
                        cursor = conn.execute("SELECT hash FROM snapshots WHERE file_path = ?", (str(full_path),))
                        row = cursor.fetchone()
                        
                        if not row:
                            changes.append(f"NEW FILE: {full_path}")
                            conn.execute("INSERT INTO snapshots VALUES (?, ?, ?)", (str(full_path), file_hash, datetime.now()))
                        elif row[0] != file_hash:
                            changes.append(f"MODIFIED: {full_path}")
                            conn.execute("UPDATE snapshots SET hash = ?, last_seen = ? WHERE file_path = ?", (file_hash, datetime.now(), str(full_path)))
                    except: pass
            
            # Check for Deleted
            cursor = conn.execute("SELECT file_path FROM snapshots")
            search_prefix = str(Path(directory))
            for (saved_path,) in cursor.fetchall():
                if saved_path.startswith(search_prefix) and saved_path not in current_files:
                    changes.append(f"DELETED: {saved_path}")
                    conn.execute("DELETE FROM snapshots WHERE file_path = ?", (saved_path,))
                    
        return "\n".join(changes) if changes else None

class DownstreamNotifier:
    def __init__(self, api_key=None, model_name="gemini-3-pro-preview"):
        self.api_key = api_key or os.environ.get("GEMINI_API_KEY")
        if not self.api_key:
            print("⚠ WARNING: GEMINI_API_KEY not found in environment.")
        
        self.model_name = model_name
        self.client = None
        if self.api_key and genai:
            self.client = genai.Client(api_key=self.api_key)

    def check_git_available(self, path="."):
        return (Path(path) / ".git").is_dir()

    def get_diff(self, path=".", range="HEAD~1"):
        if self.check_git_available(path):
            try:
                result = subprocess.run(["git", "diff", range], cwd=path, capture_output=True, check=True)
                return result.stdout.decode('utf-8', errors='ignore')
            except: pass
        
        print(f"ℹ Git repo not detected at {path}. Using SQLite snapshot diffing.")
        snapper = SnapshotDiff()
        return snapper.get_diff(path)

    def summarize_changes(self, diff_text, provider_name="a downstream provider"):
        if not self.client:
            return {
                "13.3.c": f"Manual Summary Extract: {diff_text[:500]}...",
                "13.3.d": "Manual review of code changes required."
            }

        prompt = f"""You are a compliance documentation specialist for EU AI Act. 
        Generate COMPLETE compliance documentation for {provider_name} based on the code changes below.
        
        Return a JSON object with ALL of the following keys filled with appropriate, professional content.
        Every field MUST have a value - no empty strings allowed.
        
        PROVIDER METADATA (use realistic defaults if not inferable):
        - "provider_name": "{provider_name}"
        - "provider_id": Provider registration number
        - "contact_person": Compliance officer name
        - "contact_email": Contact email
        - "contact_phone": Phone number
        - "address": Business address
        - "system_name": Name of the AI system
        - "version": Software version
        - "release_date": Release date (YYYY-MM-DD format)
        - "eu_database_id": EU AI Database registration ID
        
        ARTICLE 13 COMPLIANCE SECTIONS:
        - "13.3.a.1": Provider identity and legal name
        - "13.3.a.2": Authorised representative details
        - "13.3.a.3": Contact email for compliance
        - "13.3.a.4": Website or documentation URL
        - "13.3.b": Intended purpose and usage context
        - "13.3.b.i.1": Primary intended purpose description
        - "13.3.b.i.2": Deployment context (institutional, personal, etc.)
        - "13.3.b.ii.1": Performance metrics methodology
        - "13.3.b.ii.2": Accuracy metrics used
        - "13.3.b.ii.3": Performance validation approach
        - "13.3.c.1": Capabilities and limitations summary
        - "13.3.c.2": Known limitations and boundaries
        - "13.3.d.1": Human oversight requirements
        - "13.3.d.2": Control mechanisms description
        - "13.3.e.1": Expected lifetime
        - "13.3.e.2": Update mechanism
        - "13.3.e.3": Hardware requirements
        - "13.3.e.4": Maintenance measures
        - "13.3.f.1": Logging mechanism description
        - "13.3.f.2": Log data types collected
        - "13.3.f.3": Log storage format
        
        ANNEX XII (GPAI Models):
        - "XII.1.a": Intended tasks and integration types
        - "XII.1.b": Acceptable use policy and prohibited uses
        - "XII.1.c": Release date and distribution methods
        - "XII.1.d": Hardware interaction and API specifications
        - "XII.1.e": Software versions and dependencies
        - "XII.1.f": Model architecture summary (layers, parameters estimate)
        - "XII.1.g": Input/Output modalities and formats
        - "XII.1.h": Model license type
        - "XII.2.a": Integration instructions and infrastructure requirements
        - "XII.2.b": Context window size and token limits
        - "XII.2.c": Training data type and provenance summary
        
        DOWNSTREAM PROVIDER FIELDS:
        - "downstream_provider_name": Downstream provider business name
        - "downstream_registration": Registration number
        - "downstream_contact": Contact person name
        - "downstream_email": Contact email
        - "downstream_phone": Phone number
        - "downstream_address": Business address
        - "intended_ai_system": Description of downstream AI system
        - "intended_use_case": How GPAI model will be used
        - "risk_classification": "High-Risk" or "Limited Risk" or "Minimal Risk"
        - "integration_date": Planned integration date
        - "documentation_version": Version of documentation received
        
        CHECKLIST (all keys must have "Yes" or "No"):
        - "checklist": Dictionary with ALL these keys set to "Yes" or "No":
            "Art_13_1", "Art_13_2", "Art_13_3_a", "Art_13_3_b_i", "Art_13_3_b_ii",
            "Art_13_3_b_iii", "Art_13_3_b_iv", "Art_13_3_b_v", "Art_13_3_b_vi",
            "Art_13_3_b_vii", "Art_13_3_c", "Art_13_3_d", "Art_13_3_e", "Art_13_3_f",
            "XII_1_a", "XII_1_b", "XII_1_c", "XII_1_d", "XII_1_e", "XII_1_f", "XII_1_g", "XII_1_h",
            "XII_2_a", "XII_2_b", "XII_2_c", "Art_53_1_b"
        
        CODE CHANGES TO ANALYZE:
        {diff_text}
        
        Generate professional, non-technical compliance text. Be specific and complete.
        """
        try:
            response = self.client.models.generate_content(
                model=self.model_name, 
                contents=prompt,
                config={'response_mime_type': 'application/json'}
            )
            return json.loads(response.text)
        except Exception as e:
            print(f"Error generating structured summary: {e}")
            return {"13.3.d": f"Summary failed. Raw Diff: {diff_text[:200]}..."}

    def backup_template(self, template_path):
        """Creates a timestamped backup copy and a JSON metadata snapshot to prevent data loss."""
        try:
            template_path = Path(template_path)
            backup_dir = template_path.parent / "backups"
            backup_dir.mkdir(exist_ok=True)
            
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Binary Backup
            shutil.copy2(template_path, backup_dir / f"{template_path.stem}_{ts}.xlsx")
            
            # Data Snapshot (Other form to prevent corruption loss)
            wb = openpyxl.load_workbook(template_path, data_only=True)
            data = {}
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data[sheet_name] = [[cell.value for cell in row] for row in sheet.iter_rows(max_row=100)]
            
            snapshot_path = backup_dir / f"{template_path.stem}_{ts}_snapshot.json"
            snapshot_path.write_text(json.dumps(data, indent=2, default=str))
            return True
        except Exception as e:
            print(f"Backup failed: {e}")
            return False

    def update_excel_copy(self, template_path, output_path, compliance_data, provider_data):
        """
        Multi-field structural injection for full Article 13 automation across all sheets.
        """
        try:
            self.backup_template(template_path)
            wb = openpyxl.load_workbook(template_path)
            
            # 1. Define Global Metadata for injection
            headers = {
                "AI System Name": provider_data.get('system_name', 'SkyNet Sentinel'),
                "AI System Version": provider_data.get('version', 'v1.0.0'),
                "Identity of Provider": provider_data.get('name', 'N/A'),
                "Contact details of Provider": provider_data.get('email', 'N/A'),
                "Identity of EU representative": provider_data.get('representative', 'N/A'),
                "Model/System ID": provider_data.get('id', 'N/A'),
                "Version/Name": provider_data.get('system_name', provider_data.get('version', 'N/A')),
                "Date of Issue": datetime.now().strftime("%Y-%m-%d"),
                "Date Created": datetime.now().strftime("%Y-%m-%d"),
                "Last Updated": datetime.now().strftime("%Y-%m-%d"),
                "Author": "AI Compliance Automated Agent",
                "Reviewer": "Gemini 3 Pro Preview"
            }
            
            # 2. COMPREHENSIVE FIELD MAPPING (Label in template -> AI JSON key)
            # This maps EVERY known label in both templates to the corresponding AI output key
            field_map = {
                # Article 13 Provider Metadata
                "Provider Name": compliance_data.get('provider_name', provider_data.get('name', '')),
                "Identity of Provider": compliance_data.get('provider_name', provider_data.get('name', '')),
                "Provider Address": compliance_data.get('address', provider_data.get('address', 'N/A')),
                "Provider Contact Email": compliance_data.get('contact_email', provider_data.get('email', '')),
                "Contact details of Provider": compliance_data.get('contact_email', provider_data.get('email', '')),
                "Provider Registration Number": compliance_data.get('provider_id', provider_data.get('id', '')),
                "EU Database Registration ID": compliance_data.get('eu_database_id', 'PENDING'),
                "System Name": compliance_data.get('system_name', provider_data.get('system_name', '')),
                "Version/Name": compliance_data.get('version', provider_data.get('version', '')),
                "Date of Issue": datetime.now().strftime("%Y-%m-%d"),
                "Date Created": datetime.now().strftime("%Y-%m-%d"),
                "Last Updated": datetime.now().strftime("%Y-%m-%d"),
                "Author": "AI Compliance Automated Agent",
                "Reviewer": "Gemini 3 Pro Preview",
                "Identity of EU representative": compliance_data.get('13.3.a.2', provider_data.get('representative', 'N/A')),
                "Model/System ID": compliance_data.get('provider_id', provider_data.get('id', '')),
                
                # Article 13 Sections
                "Primary Intended Purpose": compliance_data.get('13.3.b.i.1', compliance_data.get('13.3.b', '')),
                "Intended Purpose": compliance_data.get('13.3.b', ''),
                "Deployment Context": compliance_data.get('13.3.b.i.2', ''),
                "Accuracy Metrics": compliance_data.get('13.3.b.ii.2', ''),
                "Performance Metrics": compliance_data.get('13.3.b.ii.1', ''),
                "Capabilities": compliance_data.get('13.3.c.1', ''),
                "Limitations": compliance_data.get('13.3.c.2', ''),
                "Planned System Changes": compliance_data.get('13.3.d.1', ''),
                "Human Oversight": compliance_data.get('13.3.d.1', ''),
                "Control Mechanisms": compliance_data.get('13.3.d.2', ''),
                "Expected Lifetime": compliance_data.get('13.3.e.1', ''),
                "Update Mechanism": compliance_data.get('13.3.e.2', ''),
                "Hardware Requirements": compliance_data.get('13.3.e.3', ''),
                "Maintenance Measures": compliance_data.get('13.3.e.4', ''),
                "Log Collection": compliance_data.get('13.3.f.1', ''),
                "Log Data Types": compliance_data.get('13.3.f.2', ''),
                "Log Storage Format": compliance_data.get('13.3.f.3', ''),
                
                # Annex XII GPAI Sections
                "Intended Tasks": compliance_data.get('XII.1.a', ''),
                "Integration Types": compliance_data.get('XII.1.a', ''),
                "Acceptable Use": compliance_data.get('XII.1.b', ''),
                "Prohibited Uses": compliance_data.get('XII.1.b', ''),
                "Release Date": compliance_data.get('XII.1.c', compliance_data.get('release_date', '')),
                "Distribution Methods": compliance_data.get('XII.1.c', ''),
                "Hardware Interaction": compliance_data.get('XII.1.d', ''),
                "API Specifications": compliance_data.get('XII.1.d', ''),
                "Software Versions": compliance_data.get('XII.1.e', ''),
                "Dependencies": compliance_data.get('XII.1.e', ''),
                "Model Architecture": compliance_data.get('XII.1.f', ''),
                "Number of Parameters": compliance_data.get('XII.1.f', ''),
                "Input Modality": compliance_data.get('XII.1.g', ''),
                "Output Modality": compliance_data.get('XII.1.g', ''),
                "Model License": compliance_data.get('XII.1.h', ''),
                "License Type": compliance_data.get('XII.1.h', ''),
                "Integration Instructions": compliance_data.get('XII.2.a', ''),
                "Infrastructure Requirements": compliance_data.get('XII.2.a', ''),
                "Context Window": compliance_data.get('XII.2.b', ''),
                "Token Limits": compliance_data.get('XII.2.b', ''),
                "Training Data Type": compliance_data.get('XII.2.c', ''),
                "Data Provenance": compliance_data.get('XII.2.c', ''),
                
                # Downstream Provider Fields
                "Downstream Provider Name": compliance_data.get('downstream_provider_name', provider_data.get('name', '')),
                "Contact Person": compliance_data.get('downstream_contact', provider_data.get('representative', '')),
                "Contact Email": compliance_data.get('downstream_email', provider_data.get('email', '')),
                "Contact Phone": compliance_data.get('downstream_phone', provider_data.get('phone', 'N/A')),
                "Address": compliance_data.get('downstream_address', 'N/A'),
                "Intended AI System": compliance_data.get('intended_ai_system', provider_data.get('system_name', '')),
                "Intended Use Case": compliance_data.get('intended_use_case', compliance_data.get('XII.1.a', '')),
                "Risk Classification": compliance_data.get('risk_classification', 'Limited Risk'),
                "Integration Date": compliance_data.get('integration_date', datetime.now().strftime("%Y-%m-%d")),
                "Documentation Version": compliance_data.get('documentation_version', provider_data.get('version', '1.0')),
            }

            # 3. Checklist Mapping (normalize keys for matching)
            checklist_map = compliance_data.get('checklist', {})
            
            # 4. UNIVERSAL INJECTION LOOP
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows(min_row=1, max_row=200):
                    for cell in row:
                        if cell.value:
                            val_str = str(cell.value).strip()
                            
                            # A. Field Value Injection (match label -> fill adjacent cell)
                            for label, fill_value in field_map.items():
                                if label.lower() in val_str.lower() and fill_value:
                                    # Target is the next column (or column 2 for metadata)
                                    if sheet_name == 'Document Metadata':
                                        target_col = 2
                                    else:
                                        target_col = cell.column + 1
                                        if target_col > 6: target_col = 4  # Cap at Column F
                                    
                                    try:
                                        target_cell = sheet.cell(row=cell.row, column=target_col)
                                        if not target_cell.value or str(target_cell.value).strip() == '':
                                            target_cell.value = fill_value
                                    except AttributeError:
                                        # Skip merged cells
                                        pass
                            
                            # B. Checklist Updates (match ID in Column A -> update Column C)
                            if cell.column == 1:
                                for checklist_id, status in checklist_map.items():
                                    # Normalize: XII_1_a -> XII(1)(a), Art_13_1 -> Art. 13(1)
                                    normalized_id = checklist_id.replace('_', '(').replace('(', '.', 1).replace('.', '(', 1) if '_' in checklist_id else checklist_id
                                    if checklist_id in val_str or normalized_id in val_str:
                                        target_cell = sheet.cell(row=cell.row, column=3)
                                        val_curr = str(target_cell.value) if target_cell.value else ""
                                        
                                        if "☐ Yes / ☐ No" in val_curr or "☐ Yes" in val_curr:
                                            if str(status).lower() == 'yes':
                                                target_cell.value = "☑ Yes / ☐ No"
                                            else:
                                                target_cell.value = "☐ Yes / ☑ No"
                                        elif "☐ Complete" in val_curr:
                                            target_cell.value = "☑ Complete"
                                        break
            
            wb.save(output_path)
            wb.close()
            return True
        except Exception as e:
            print(f" [ERROR] Excel Update Failed: {e}")
            return False

    def read_annex_xii_input(self, examples_dir):
        """Read Annex XII JSON files as INPUT (GPAI model info provided by user/fake data)."""
        annex_data = []
        for file in Path(examples_dir).glob("*_AnnexXII.json"):
            try:
                data = json.loads(file.read_text())
                data['source_file'] = file.name
                annex_data.append(data)
            except Exception as e:
                print(f"  Warning: Failed to read {file.name}: {e}")
        return annex_data
    
    def read_providers(self, examples_dir):
        """Read provider registry (for email routing). Falls back to Annex XII data."""
        providers = []
        # Try Excel provider files first
        for file in Path(examples_dir).glob("*_Provider.xlsx"):
            try:
                df = pd.read_excel(file)
                providers.extend(df.to_dict('records'))
            except: pass
        
        # If no Excel providers, use Annex XII JSON as provider source
        if not providers:
            for file in Path(examples_dir).glob("*_AnnexXII.json"):
                try:
                    data = json.loads(file.read_text())
                    providers.append({
                        'name': data.get('provider_name', 'Unknown'),
                        'id': data.get('provider_id', 'N/A'),
                        'email': data.get('contact_email', 'N/A'),
                        'version': data.get('model_version', 'v1.0'),
                        'system_name': data.get('model_name', 'AI System'),
                        'annex_xii': data  # Include full Annex XII data
                    })
                except: pass
        return providers

    def create_json_manifest(self, output_dir, provider_data, summary, files):
        recipient = provider_data.get('email', 'N/A')
        manifest = {
            "metadata": {
                "to": recipient,
                "subject": "Article 13 Compliance Update",
                "timestamp": datetime.now().isoformat()
            },
            "content": {
                "body": f"Dear {provider_data.get('name', 'Provider')},\n\nPlease see attached compliance updates.",
                "summary_text": summary
            },
            "attachments": [str(f.resolve()) for f in files]
        }
        manifest_path = output_dir / f"Notice_{provider_data.get('id', 'Unknown')}.json"
        manifest_path.write_text(json.dumps(manifest, indent=2))
        return manifest_path

    def run_md_to_pdf(self, md_path, output_path):
        script_dir = Path(__file__).resolve().parent
        pdf_tool = script_dir.parent.parent.parent.parent / "skills" / "explaining-code" / "md_to_pdf.py"
        if not pdf_tool.exists(): pdf_tool = script_dir.parent.parent.parent / "md_to_pdf.py"
        if pdf_tool.exists():
            try:
                subprocess.run([sys.executable, str(pdf_tool), str(md_path), str(output_path)], capture_output=True)
                return True
            except: return False
        return False

    def process(self, repo_path, examples_dir, templates_dir, manual_text=None):
        print(f"[START] Scanning {repo_path}")
        
        diff = manual_text if manual_text else self.get_diff(repo_path)
        if not diff:
            print("No changes detected. Skipping notification.")
            return

        providers = self.read_providers(examples_dir)
        if not providers:
            print(f"No providers found in {examples_dir}")
            return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        batch_dir = Path("Output") / f"Batch_{ts}"
        batch_dir.mkdir(parents=True, exist_ok=True)
        
        # Identify templates
        templates = list(Path(templates_dir).glob("*.xlsx"))
        templates = [t for t in templates if "backup" not in str(t).lower() and "~$" not in t.name]

        for p in providers:
            pname = p.get('name', 'Provider')
            pid = p.get('id', 'P00')
            
            # Get Annex XII data (INPUT) if available
            annex_xii = p.get('annex_xii', {})
            
            # Generate structured compliance data personalized for this provider
            compliance_data_raw = self.summarize_changes(diff, provider_name=pname)
            
            # Flatten the nested dictionary (AI returns grouped sections)
            compliance_data = {}
            for k, v in compliance_data_raw.items():
                if isinstance(v, dict):
                    compliance_data.update(v)
                else:
                    compliance_data[k] = v
            
            # Merge Annex XII INPUT into compliance data
            # This ensures Article 13 OUTPUT can reference GPAI model info
            if annex_xii:
                compliance_data.update({
                    'XII.1.a': annex_xii.get('intended_tasks', compliance_data.get('XII.1.a', '')),
                    'XII.1.b': annex_xii.get('prohibited_uses', compliance_data.get('XII.1.b', '')),
                    'XII.1.c': annex_xii.get('release_date', compliance_data.get('XII.1.c', '')),
                    'XII.1.d': annex_xii.get('hardware_requirements', compliance_data.get('XII.1.d', '')),
                    'XII.1.e': annex_xii.get('software_versions', compliance_data.get('XII.1.e', '')),
                    'XII.1.f': annex_xii.get('architecture', compliance_data.get('XII.1.f', '')),
                    'XII.1.g': annex_xii.get('input_modality', compliance_data.get('XII.1.g', '')),
                    'XII.1.h': annex_xii.get('license', compliance_data.get('XII.1.h', '')),
                    'XII.2.a': annex_xii.get('integration_instructions', compliance_data.get('XII.2.a', '')),
                    'XII.2.b': annex_xii.get('context_window', compliance_data.get('XII.2.b', '')),
                    'XII.2.c': annex_xii.get('training_data_type', compliance_data.get('XII.2.c', '')),
                    # Provider metadata from Annex XII
                    'provider_name': annex_xii.get('provider_name', pname),
                    'provider_id': annex_xii.get('provider_id', pid),
                    'contact_email': annex_xii.get('contact_email', ''),
                    'model_name': annex_xii.get('model_name', ''),
                    'model_version': annex_xii.get('model_version', ''),
                })
                # Merge checklist
                if 'checklist' in annex_xii:
                    if 'checklist' not in compliance_data:
                        compliance_data['checklist'] = {}
                    compliance_data['checklist'].update(annex_xii['checklist'])

            print(f"DEBUG: Keys: {list(compliance_data.keys())}")
            
            # GENERIC MD TEXT (User Request)
            md_text = f"""# Regulatory Compliance Notification
**Provider:** {pname} ({pid})
**Date:** {datetime.now().strftime("%Y-%m-%d")}

Please refer to the attached Excel documentation for full Article 13 and Annex XII compliance details. 
This summary is a notification of changes only."""
            
            # Process ALL templates -> One output per template per provider
            generated_files = []
            for tmpl in templates:
                base_name = tmpl.stem.replace("_Template", "").replace("Template", "")
                out_name = f"{base_name}_{pid}.xlsx"
                excel_out = batch_dir / out_name
                
                print(f"   Generating {out_name} from {tmpl.name}...")
                if self.update_excel_copy(tmpl, excel_out, compliance_data, p):
                    generated_files.append(excel_out)

            md_out = batch_dir / f"Summary_{pid}.md"
            md_out.write_text(md_text, encoding='utf-8')
            pdf_out = batch_dir / f"Summary_{pid}.pdf"
            self.run_md_to_pdf(md_out, pdf_out)
            generated_files.append(pdf_out)
            
            self.create_json_manifest(batch_dir, p, md_text, generated_files)
            print(f"[DONE] Hardened Package ready for {pname} (ID: {pid})")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--repo", default=".")
    parser.add_argument("--examples", default="examples")
    parser.add_argument("--templates", default="templates") # Changed from single file to dir
    parser.add_argument("--manual")
    # Placeholder for snapshot_mode if it exists elsewhere in the full code
    parser.add_argument("--snapshot_mode", action="store_true", help="Enable snapshot mode (if applicable)") 
    args = parser.parse_args()

    notifier = DownstreamNotifier()
    if args.snapshot_mode:
       # ... logic remains or needs minor tweak if snapshot mode uses single file
       pass 
    else:
        notifier.process(args.repo, args.examples, args.templates, args.manual)

if __name__ == "__main__":
    main()
